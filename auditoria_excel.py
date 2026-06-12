from __future__ import annotations

import json
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

MAX_FILAS_EXCEL = 1_048_000

AZUL = "1F4E79"
AZUL_CLARO = "D5E8F0"
GRIS_BORDE = "BFBFBF"

DESCRIPCIONES: dict[str, str] = {
    "casos_duplicados":           "Casos con mensajes sustantivos duplicados. Los registros repetidos se eliminaron; saludos repetidos no marcan esta categoría.",
    "casos_llamadas":             "Casos atendidos por LLAMADA: el chat solo no forma conversación. Excluidos.",
    "casos_con_llamada":          "Casos con marcador de llamada PERO con chat completo. Sí se calcularon.",
    "casos_solo_inbound":         "2+ mensajes, todos del cliente, sin respuesta registrada. Excluidos. (Los de un solo mensaje se descartan sin marcar.)",
    "casos_solo_outbound":        "2+ mensajes, todos salientes; el cliente nunca escribió. Excluidos. (Los de un solo mensaje se descartan sin marcar.)",
    "casos_mensaje_repetido":     "El cliente envió 2+ veces exactamente el mismo mensaje y nada más: no es conversación. Excluidos.",
    "casos_fechas_incoherentes":  "Un mismo nro_caso con conversaciones separadas por un hueco > 24h (fechas que no cuadran). Excluidos.",
    "casos_colaborador":          "Sin palabras de venta y con vocabulario coloquial/interno: posible charla entre colaboradores. Excluidos.",
    "casos_sin_atencion_humana":  "Conversaciones SIN respuesta de agente humano posterior al cliente. Excluidos.",
    "casos_sin_bot":              "Conversaciones atendidas por humano SIN paso previo de bot. Sí se calcularon (incluidas en métricas); se listan aparte.",
    "casos_extraordinarios":      "Casos con más de 30 mensajes (conversación inusualmente larga). Sí se calcularon; revisar.",
    "casos_sin_palabras_venta":   "Conversación sin vocabulario de venta (posible charla interna). Sí se calcularon.",
    "casos_vacios":               "Sin mensajes válidos tras la limpieza. Excluidos.",
}

_borde = Side(style="thin", color=GRIS_BORDE)
BORDES = Border(left=_borde, right=_borde, top=_borde, bottom=_borde)
FUENTE = Font(name="Arial", size=10)
FUENTE_BOLD = Font(name="Arial", size=10, bold=True)
FUENTE_HEADER = Font(name="Arial", size=10, bold=True, color="FFFFFF")
FILL_HEADER = PatternFill("solid", start_color=AZUL)
FILL_SUAVE = PatternFill("solid", start_color=AZUL_CLARO)
CENTRO = Alignment(horizontal="center")


def _celda(ws, fila, col, valor, font=FUENTE, fill=None, borde=True,
           align=None, numfmt=None):
    c = ws.cell(row=fila, column=col, value=valor)
    c.font = font
    if fill:
        c.fill = fill
    if borde:
        c.border = BORDES
    if align:
        c.alignment = align
    if numfmt:
        c.number_format = numfmt
    return c


def _caso_fecha(c) -> tuple:
    if isinstance(c, dict):
        return c.get("nro_caso"), c.get("fecha", "")
    return c, ""


def _orden_categorias(cats) -> list[str]:
    orden = [c for c in DESCRIPCIONES if c in cats]
    return orden + [c for c in cats if c not in DESCRIPCIONES]


def _titulo(ws, texto, sub1, sub2):
    ws.sheet_view.showGridLines = False
    _celda(ws, 1, 1, texto,
           font=Font(name="Arial", size=14, bold=True, color=AZUL), borde=False)
    _celda(ws, 2, 1, sub1, borde=False)
    _celda(ws, 3, 1, sub2, borde=False)


# 1) EXCEL POR PERIODO
def escribir_auditoria_excel(auditoria: dict, ruta_xlsx: Path | str,
                             fecha_desde: date | None = None,
                             fecha_hasta: date | None = None) -> Path:
    ruta_xlsx = Path(ruta_xlsx)
    categorias = _orden_categorias(auditoria.keys())

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"
    periodo = (f"Periodo: {fecha_desde} a {fecha_hasta}"
               if fecha_desde and fecha_hasta else "Periodo: (no indicado)")
    _titulo(ws, "Auditoría de casos — WhatsApp ETL", periodo,
            f"Generado: {datetime.now():%Y-%m-%d %H:%M}")

    fila0 = 5
    for j, h in enumerate(("Categoría", "N° de casos", "Descripción"), start=1):
        _celda(ws, fila0, j, h, font=FUENTE_HEADER, fill=FILL_HEADER, align=CENTRO)
    fila = fila0 + 1
    for cat in categorias:
        _celda(ws, fila, 1, cat)
        _celda(ws, fila, 2, f"=COUNTA('{cat}'!A:A)-1", align=CENTRO)
        _celda(ws, fila, 3, DESCRIPCIONES.get(cat, ""))
        fila += 1
    _celda(ws, fila, 1, "TOTAL", font=FUENTE_BOLD, fill=FILL_SUAVE)
    _celda(ws, fila, 2, f"=SUM(B{fila0 + 1}:B{fila - 1})",
           font=FUENTE_BOLD, fill=FILL_SUAVE, align=CENTRO)
    _celda(ws, fila, 3, "Apariciones en categorías (un caso puede estar en varias).",
           fill=FILL_SUAVE)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 95
    ws.freeze_panes = f"A{fila0 + 1}"

    ws_t = wb.create_sheet("Todos")
    for j, h in enumerate(("nro_caso", "fecha", "categoria"), start=1):
        _celda(ws_t, 1, j, h, font=FUENTE_HEADER, fill=FILL_HEADER)
    f = 2
    for cat in categorias:
        for caso in auditoria.get(cat, []):
            nro, fch = _caso_fecha(caso)
            _celda(ws_t, f, 1, nro)
            _celda(ws_t, f, 2, fch)
            _celda(ws_t, f, 3, cat)
            f += 1
    ws_t.auto_filter.ref = f"A1:C{max(f - 1, 1)}"
    ws_t.freeze_panes = "A2"
    ws_t.column_dimensions["A"].width = 14
    ws_t.column_dimensions["B"].width = 18
    ws_t.column_dimensions["C"].width = 28

    for cat in categorias:
        wsc = wb.create_sheet(cat[:31])
        _celda(wsc, 1, 1, "nro_caso", font=FUENTE_HEADER, fill=FILL_HEADER)
        _celda(wsc, 1, 2, "fecha", font=FUENTE_HEADER, fill=FILL_HEADER)
        casos = auditoria.get(cat, [])
        for i, caso in enumerate(casos, start=2):
            nro, fch = _caso_fecha(caso)
            _celda(wsc, i, 1, nro)
            _celda(wsc, i, 2, fch)
        wsc.auto_filter.ref = f"A1:B{max(len(casos) + 1, 1)}"
        wsc.freeze_panes = "A2"
        wsc.column_dimensions["A"].width = 14
        wsc.column_dimensions["B"].width = 18

    ruta_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(ruta_xlsx)
    return ruta_xlsx


# 2) EXCEL GENERAL ACUMULADO
def actualizar_auditoria_general(auditoria: dict, fecha_desde: date,
                                 fecha_hasta: date,
                                 output_dir: Path | str = "output/auditoria") -> Path | None:
    """Agrega/reemplaza el periodo en el maestro JSON y regenera el Excel
    general. Devuelve la ruta del xlsx, o None si estaba abierto en Excel
    (el JSON queda actualizado igual; regenerar luego con --general)."""
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    ruta_json = output_dir / "auditoria_general.json"

    master: dict = {}
    if ruta_json.exists():
        master = json.loads(ruta_json.read_text(encoding="utf-8"))
    clave = f"{fecha_desde.isoformat()}_{fecha_hasta.isoformat()}"
    master[clave] = auditoria  # mismo periodo -> se reemplaza (no duplica)
    ruta_json.write_text(json.dumps(master, ensure_ascii=False, indent=2),
                         encoding="utf-8")

    ruta_xlsx = output_dir / "auditoria_general.xlsx"
    try:
        _escribir_general(master, ruta_xlsx)
    except PermissionError:
        return None
    return ruta_xlsx


def _escribir_general(master: dict, ruta_xlsx: Path) -> None:
    periodos = sorted(master.keys())
    periodos_detalle = sorted(master.keys(), reverse=True)
    etiquetas = {p: p.replace("_", " a ") for p in periodos}
    cats: list[str] = []
    for aud in master.values():
        for c in aud:
            if c not in cats:
                cats.append(c)
    categorias = _orden_categorias(cats)

    wb = Workbook()

    # Resumen
    ws = wb.active
    ws.title = "Resumen"
    rango = (f"Periodos acumulados: {len(periodos)} "
             f"({etiquetas[periodos[0]]} … {etiquetas[periodos[-1]]})"
             if periodos else "Sin periodos aún")
    _titulo(ws, "Auditoría GENERAL de casos — WhatsApp ETL", rango,
            f"Actualizado: {datetime.now():%Y-%m-%d %H:%M}")
    fila0 = 5
    for j, h in enumerate(("Categoría", "Total casos", "Descripción"), start=1):
        _celda(ws, fila0, j, h, font=FUENTE_HEADER, fill=FILL_HEADER, align=CENTRO)
    fila = fila0 + 1
    for cat in categorias:
        _celda(ws, fila, 1, cat)
        _celda(ws, fila, 2,
               sum(len(master[p].get(cat, [])) for p in periodos),
               align=CENTRO)
        _celda(ws, fila, 3, DESCRIPCIONES.get(cat, ""))
        fila += 1
    _celda(ws, fila, 1, "TOTAL", font=FUENTE_BOLD, fill=FILL_SUAVE)
    _celda(ws, fila, 2, f"=SUM(B{fila0 + 1}:B{fila - 1})",
           font=FUENTE_BOLD, fill=FILL_SUAVE, align=CENTRO)
    _celda(ws, fila, 3, "Apariciones acumuladas en todos los periodos.",
           fill=FILL_SUAVE)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 95
    ws.freeze_panes = f"A{fila0 + 1}"

    # Por periodo
    ws_p = wb.create_sheet("Por periodo")
    _celda(ws_p, 1, 1, "Periodo", font=FUENTE_HEADER, fill=FILL_HEADER)
    for j, cat in enumerate(categorias, start=2):
        _celda(ws_p, 1, j, cat, font=FUENTE_HEADER, fill=FILL_HEADER, align=CENTRO)
    col_total = len(categorias) + 2
    _celda(ws_p, 1, col_total, "TOTAL", font=FUENTE_HEADER, fill=FILL_HEADER,
           align=CENTRO)
    for i, p in enumerate(periodos, start=2):
        _celda(ws_p, i, 1, etiquetas[p])
        for j, cat in enumerate(categorias, start=2):
            _celda(ws_p, i, j, len(master[p].get(cat, [])), align=CENTRO)
        _celda(ws_p, i, col_total,
               f"=SUM(B{i}:{get_column_letter(col_total - 1)}{i})",
               font=FUENTE_BOLD, fill=FILL_SUAVE, align=CENTRO)
    ws_p.column_dimensions["A"].width = 26
    for j in range(2, col_total + 1):
        ws_p.column_dimensions[get_column_letter(j)].width = 19
    ws_p.freeze_panes = "B2"

    # Todos los casos juntos
    ws_t = wb.create_sheet("Todos")
    for j, h in enumerate(("nro_caso", "fecha", "categoria", "periodo"), start=1):
        _celda(ws_t, 1, j, h, font=FUENTE_HEADER, fill=FILL_HEADER)
    f = 2
    truncado = False
    for p in periodos_detalle:
        aud = master[p]
        for cat in categorias:
            for caso in aud.get(cat, []):
                if f > MAX_FILAS_EXCEL:
                    truncado = True
                    break
                nro, fch = _caso_fecha(caso)
                _celda(ws_t, f, 1, nro)
                _celda(ws_t, f, 2, fch)
                _celda(ws_t, f, 3, cat)
                _celda(ws_t, f, 4, etiquetas[p])
                f += 1
            if truncado:
                break
        if truncado:
            break
    if truncado:
        _celda(ws_t, f, 1, "LISTA TRUNCADA (límite de Excel): se muestra de lo "
               "más reciente a lo más antiguo; detalle completo en los "
               "auditoria_<periodo>.csv", font=FUENTE_BOLD, fill=FILL_SUAVE)
    ws_t.auto_filter.ref = f"A1:D{max(f - 1, 1)}"
    ws_t.freeze_panes = "A2"
    ws_t.column_dimensions["A"].width = 14
    ws_t.column_dimensions["B"].width = 18
    ws_t.column_dimensions["C"].width = 28
    ws_t.column_dimensions["D"].width = 26

    # Una hoja por categoría
    for cat in categorias:
        wsc = wb.create_sheet(cat[:31])
        for j, h in enumerate(("nro_caso", "fecha", "periodo"), start=1):
            _celda(wsc, 1, j, h, font=FUENTE_HEADER, fill=FILL_HEADER)
        i = 2
        trunc_cat = False
        for p in periodos_detalle:
            for caso in master[p].get(cat, []):
                if i > MAX_FILAS_EXCEL:
                    trunc_cat = True
                    break
                nro, fch = _caso_fecha(caso)
                _celda(wsc, i, 1, nro)
                _celda(wsc, i, 2, fch)
                _celda(wsc, i, 3, etiquetas[p])
                i += 1
            if trunc_cat:
                break
        if trunc_cat:
            _celda(wsc, i, 1, "LISTA TRUNCADA (límite de Excel)",
                   font=FUENTE_BOLD, fill=FILL_SUAVE)
        wsc.auto_filter.ref = f"A1:C{max(i - 1, 1)}"
        wsc.freeze_panes = "A2"
        wsc.column_dimensions["A"].width = 14
        wsc.column_dimensions["B"].width = 18
        wsc.column_dimensions["C"].width = 26

    ruta_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(ruta_xlsx)


def main() -> None:
    import argparse

    ap = argparse.ArgumentParser(description="Excel de auditoría (periodo o general).")
    ap.add_argument("--fecha", type=str, help="YYYY-MM-DD (un solo día)")
    ap.add_argument("--desde", type=str)
    ap.add_argument("--hasta", type=str)
    ap.add_argument("--general", action="store_true",
                    help="Regenerar auditoria_general.xlsx desde el JSON maestro")
    ap.add_argument("--listar-periodos", action="store_true",
                    help="Mostrar los periodos acumulados en el maestro")
    ap.add_argument("--quitar-periodo", type=str, metavar="CLAVE",
                    help="Eliminar un periodo del maestro (ej: 2026-04-01_2026-04-02)")
    args = ap.parse_args()

    ruta_master = Path("output/auditoria/auditoria_general.json")

    if args.listar_periodos:
        if not ruta_master.exists():
            print("No existe el maestro todavía.")
            return
        master = json.loads(ruta_master.read_text(encoding="utf-8"))
        for clave in sorted(master):
            n = sum(len(v) for v in master[clave].values())
            print(f"  {clave}   ({n} apariciones)")
        return

    if args.quitar_periodo:
        if not ruta_master.exists():
            print("No existe el maestro todavía.")
            return
        master = json.loads(ruta_master.read_text(encoding="utf-8"))
        if args.quitar_periodo not in master:
            print(f"No existe el periodo '{args.quitar_periodo}'. "
                  "Usa --listar-periodos para ver las claves.")
            return
        del master[args.quitar_periodo]
        ruta_master.write_text(json.dumps(master, ensure_ascii=False, indent=2),
                               encoding="utf-8")
        try:
            _escribir_general(master, Path("output/auditoria/auditoria_general.xlsx"))
            print(f"Periodo '{args.quitar_periodo}' eliminado y Excel regenerado.")
        except PermissionError:
            print("Periodo eliminado del JSON. Cierra el Excel y corre --general.")
        return

    if args.general:
        ruta_json = Path("output/auditoria/auditoria_general.json")
        if not ruta_json.exists():
            print("No existe output/auditoria_general.json todavía.")
            return
        master = json.loads(ruta_json.read_text(encoding="utf-8"))
        try:
            _escribir_general(master, Path("output/auditoria/auditoria_general.xlsx"))
            print("OK -> output/auditoria_general.xlsx")
        except PermissionError:
            print("Cierra auditoria_general.xlsx en Excel y vuelve a intentar.")
        return

    if args.fecha:
        fd = fh = datetime.strptime(args.fecha, "%Y-%m-%d").date()
    elif args.desde and args.hasta:
        fd = datetime.strptime(args.desde, "%Y-%m-%d").date()
        fh = datetime.strptime(args.hasta, "%Y-%m-%d").date()
    else:
        print("Indica --fecha, --desde/--hasta, o --general.")
        return

    base = f"auditoria_{fd.isoformat()}_{fh.isoformat()}"
    ruta_json = Path("output/auditoria") / f"{base}.json"
    if not ruta_json.exists():
        print(f"No existe {ruta_json}. Corre primero --solo-metricas.")
        return
    auditoria = json.loads(ruta_json.read_text(encoding="utf-8"))
    out = escribir_auditoria_excel(auditoria, Path("output/auditoria") / f"{base}.xlsx", fd, fh)
    print(f"OK -> {out}")
    # y de paso lo suma al general
    res = actualizar_auditoria_general(auditoria, fd, fh)
    print(f"OK -> {res}" if res else
          "auditoria_general.xlsx estaba abierto; JSON actualizado. "
          "Regenera con: python auditoria_excel.py --general")


if __name__ == "__main__":
    main()